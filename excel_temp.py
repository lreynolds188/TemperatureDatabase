from openpyxl import *
from openpyxl.chart import *
from openpyxl.chart.axis import DateAxis
import utilities
import sys

conn = utilities.ConnectDatabase()
cursor = conn.cursor()
wb_filename = 'assets\World Temperature.xlsx'
worldTemp_workbook = Workbook()
worldTemp_worksheet = worldTemp_workbook.active
worldTemp_lineChart = LineChart()


def CreateWorkbook():
    try:
        sys.stdout.write('Creating World_Temperature workbook... ')
        sys.stdout.flush()

        worldTemp_worksheet = worldTemp_workbook.active
        worldTemp_worksheet.title = "Temperature by City"

        print('World_Temperature workbook created successfully.\n')
    except:
        print('Error E1: Failed to load World_Temperature workbook.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def CreateLineChart():
    try:
        sys.stdout.write('Creating World_Temperature LineChart... ')
        sys.stdout.flush()

        # setup linechart
        worldTemp_lineChart.title = "China Temperature by Date"
        worldTemp_lineChart.style = 12
        worldTemp_lineChart.y_axis.title = 'Temp'
        worldTemp_lineChart.x_axis.title = 'Date'
        worldTemp_lineChart.x_axis.tickLblPos = "low"

        data = Reference(worldTemp_worksheet, min_col=2, min_row=1, max_col=2, max_row=2200)
        worldTemp_lineChart.add_data(data, titles_from_data=True)
        dates = Reference(worldTemp_worksheet, min_col=1, min_row=1, max_col=1, max_row=2200)
        worldTemp_lineChart.set_categories(dates)
        worldTemp_worksheet.add_chart(worldTemp_lineChart, "C1")

        print('Creation successful.\n')
    except:
        print('Error E2: LineChart creation failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def SelectChinaTempData():
    try:
        sys.stdout.write('Selecting data from Temperatures_by_Country... ')
        sys.stdout.flush()

        _query = "SELECT SUBSTR(EventDate, 0, 5) as Year, AverageTemperature FROM GlobalLandTemperaturesByCountry WHERE Country = 'China' AND AverageTemperature != 'NULL' ORDER BY Year"
        cursor.execute(_query)

        print('Data select successful.\n')
    except:
        print('Error E3: Data select failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def InsertDataIntoWorkbook():
    try:
        sys.stdout.write('Inserting data into World_Temperature workbook... ')
        sys.stdout.flush()

        _chinaTempData = cursor.fetchall()

        for rows in _chinaTempData:
            worldTemp_worksheet.append(rows)

        print('Data successfully inserted.\n')
    except:
        print('Error E4: Data insert failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def Run():
    CreateWorkbook()
    SelectChinaTempData()
    InsertDataIntoWorkbook()
    CreateLineChart()
    worldTemp_workbook.save(wb_filename)
    utilities.CloseDatabaseConnection(conn)


CreateWorkbook()
SelectChinaTempData()
InsertDataIntoWorkbook()
CreateLineChart()
worldTemp_workbook.save(wb_filename)
utilities.CloseDatabaseConnection(conn)
