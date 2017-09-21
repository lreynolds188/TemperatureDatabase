from openpyxl import load_workbook
import utilities
import sys

conn = utilities.ConnectDatabase()
cursor = conn.cursor()


def CreateCountryWorkbook():
    filename_country = 'assets/GlobalLandTemperaturesByCountry.xlsx'

    # Workbook 1
    sys.stdout.write('Loading Country workbook... ')
    sys.stdout.flush()
    try:
        workbook_country = load_workbook(filename_country)
        sheet_ranges_country = workbook_country.sheetnames
        sheet_country = workbook_country[sheet_ranges_country[0]]
        print('Workbook load successful.\n')
        return sheet_country;
    except:
        print('Error C1: Failed to load Country workbook.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def CreateMajorCityWorkbook():
    filename_major_city = 'assets/GlobalLandTemperaturesByMajorCity.xlsx'

    # Workbook 2
    sys.stdout.write('Loading Major_City workbook... ')
    sys.stdout.flush()
    try:
        workbook_major_city = load_workbook(filename_major_city)
        sheet_ranges_major_city = workbook_major_city.sheetnames
        sheet_major_city = workbook_major_city[sheet_ranges_major_city[0]]
        print('Workbook load successful.\n')
        return sheet_major_city;
    except:
        print('Error C2: Failed to load Major_City workbook.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def CreateStateWorkbook():
    filename_state = 'assets/GlobalLandTemperaturesByState.xlsx'

    sys.stdout.write('Loading State workbook... ')
    sys.stdout.flush()
    try:
        workbook_state = load_workbook(filename_state)
        sheet_ranges_state = workbook_state.sheetnames
        sheet_state = workbook_state[sheet_ranges_state[0]]
        print('Workbook load successful.\n')
        return sheet_state;
    except:
        print('Error C3: Failed to load State workbook.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def CreateDatabase():
    try:
        sys.stdout.write('Creating database... ')
        sys.stdout.flush()
        cursor.execute('''DROP TABLE IF EXISTS GlobalLandTemperaturesByCountry''')
        cursor.execute('''CREATE TABLE GlobalLandTemperaturesByCountry (EventDate VARCHAR(25), AverageTemperature DECIMAL(25), AverageTemperatureUncertainty DECIMAL(25), Country VARCHAR(25))''')
        cursor.execute('''DROP TABLE IF EXISTS GlobalLandTemperaturesByMajorCity''')
        cursor.execute('''CREATE TABLE GlobalLandTemperaturesByMajorCity (EventDate VARCHAR(25), AverageTemperature DECIMAL(25), AverageTemperatureUncertainty DECIMAL(25), City VARCHAR(25), Country VARCHAR(25), Latitude VARCHAR(25), Longitude VARCHAR(25))''')
        cursor.execute('''DROP TABLE IF EXISTS GlobalLandTemperaturesByState''')
        cursor.execute('''CREATE TABLE GlobalLandTemperaturesByState (EventDate VARCHAR(25), AverageTemperature DECIMAL(25), AverageTemperatureUncertainty DECIMAL(25), State VARCHAR(25), Country VARCHAR(25))''')
        conn.commit()
        print('Database creation successful.\n')
    except:
        print('Error C4: Database creation failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def InsertCountryWorksheet(_worksheet):
    try:
        sys.stdout.write('Inserting Country data into database... ')
        sys.stdout.flush()
        for row in _worksheet.iter_rows(min_row=2):
            _date = row[0].value
            _averageTemp = row[1].value
            _averageTempUncertain = row[2].value
            _country = row[3].value
            _query = "INSERT INTO GlobalLandTemperaturesByCountry (EventDate, AverageTemperature, AverageTemperatureUncertainty, Country) VALUES (?, ?, ?, ?)"
            _values = (_date, _averageTemp, _averageTempUncertain, _country)

            cursor.execute(_query, _values)
        conn.commit()
        print('Data insert successful.\n')
    except:
        print('Error C5: Data insert failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def InsertMajorCityWorksheet(_worksheet):
    try:
        sys.stdout.write('Inserting Major_City data into database... ')
        sys.stdout.flush()
        for row in _worksheet.iter_rows(min_row=2):
            _date = row[0].value
            _averageTemp = row[1].value
            _averageTempUncertain = row[2].value
            _city = row[3].value
            _country = row[4].value
            _latitude = row[5].value
            _longitude = row[6].value

            _query = "INSERT INTO GlobalLandTemperaturesByMajorCity (EventDate, AverageTemperature, AverageTemperatureUncertainty, City, Country, Latitude, Longitude) VALUES (?, ?, ?, ?, ?, ?, ?)"
            _values = (_date, _averageTemp, _averageTempUncertain, _city, _country, _latitude, _longitude)

            cursor.execute(_query, _values)
        conn.commit()
        print('Data insert successful.\n')
    except:
        print('Error C6: Database connection failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def InsertStateWorksheet(_worksheet):
    try:
        sys.stdout.write('Inserting State data into database... ')
        sys.stdout.flush()
        for row in _worksheet.iter_rows(min_row=2):
            _date = row[0].value
            _averageTemp = row[1].value
            _averageTempUncertain = row[2].value
            _state = row[3].value
            _country = row[4].value

            _query = "INSERT INTO GlobalLandTemperaturesByState (EventDate, AverageTemperature, AverageTemperatureUncertainty, State, Country) VALUES (?, ?, ?, ?, ?)"
            _values = (_date, _averageTemp, _averageTempUncertain, _state, _country)

            cursor.execute(_query, _values)
        conn.commit()
        print('Data insert successful.\n')
    except:
        print('Error C7: State data insert failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


CreateDatabase()

country_worksheet = CreateCountryWorkbook()
majorCity_worksheet = CreateMajorCityWorkbook()
state_worksheet = CreateStateWorkbook()

InsertCountryWorksheet(country_worksheet)
InsertMajorCityWorksheet(majorCity_worksheet)
InsertStateWorksheet(state_worksheet)

utilities.CloseDatabaseConnection(conn)