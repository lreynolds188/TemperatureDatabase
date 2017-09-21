from openpyxl import load_workbook
from openpyxl import Workbook
import utilities
import sys
import matplotlib.pyplot as pyplot
from matplotlib.legend_handler import HandlerLine2D
import numpy

conn = utilities.ConnectDatabase()
cursor = conn.cursor()
wb_filename = 'assets\World Temperature.xlsx'
worldTemp_workbook = Workbook()
actTempArr, nswTempArr, ntTempArr, qldTempArr, saTempArr, tasTempArr, vicTempArr, waTempArr, ausTempArr = [], [], [], [], [], [], [], [], []
actYearArr, nswYearArr, ntYearArr, qldYearArr, saYearArr, tasYearArr, vicYearArr, waYearArr, ausYearArr = [], [], [], [], [], [], [], [], []


def LoadWorkbook(_filename):
    try:
        sys.stdout.write('Loading World_Temperature workbook... ')
        sys.stdout.flush()

        _workbook = load_workbook(_filename)

        print('World_Temperature workbook loaded successfully.\n')
        return _workbook
    except:
        print('Error N1: Failed to load World_Temperature workbook.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def AddWorkbookSheet(_workbook):
    try:
        DeleteWorksheet(_workbook)

        sys.stdout.write('Creating new worksheet... ')
        sys.stdout.flush()

        _worksheet = _workbook.create_sheet('Comparison')
        _worksheet.title = "Comparison"

        print('Worksheet created successfully.\n')
        return _worksheet
    except:
        print('Error N2: Failed to create worksheet.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def DeleteWorksheet(_workbook):
    try:
        sys.stdout.write('Deleting worksheet... ')
        sys.stdout.flush()

        _tempsheet = _workbook.get_sheet_by_name('Comparison')
        _workbook.remove_sheet(_tempsheet)

        print('Worksheet deleted successfully.\n')
    except:
        print('Error N3: No worksheet found.\n')


def SelectStateTempData(_query):
    try:
        sys.stdout.write('Selecting data from Temperatures_by_State... ')
        sys.stdout.flush()

        cursor.execute(_query)

        print('Data select successful.\n')
    except:
        print('Error N4: Data select failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def InsertDataIntoWorkbook(_worksheet):
    try:
        sys.stdout.write('Inserting data into World_Temperature workbook... ')
        sys.stdout.flush()

        _stateTempData = cursor.fetchall()

        for row in _stateTempData:
            _worksheet.append(row)

        print('Data successfully inserted.\n')
    except:
        print('Error N5: Data insert failed.\n')
        utilities.CloseDatabaseConnection(conn)
        exit(1)


def ImportPlotData(_worksheet):
    tempArr = []
    for row in _worksheet.iter_rows(min_row=1):
        tempRow = []
        for i in row:
            tempRow.append(i.value)
        tempArr.append((tempRow))

    for row in tempArr:
        var = row[0]
        if (var == 'Australian Capital Territory'):
            actTempArr.append(row[2])
            actYearArr.append(row[1])
        elif (var == 'New South Wales'):
            nswTempArr.append(row[2])
            nswYearArr.append(row[1])
        elif (var == 'Northern Territory'):
            ntTempArr.append(row[2])
            ntYearArr.append(row[1])
        elif (var == 'Queensland'):
            qldTempArr.append(row[2])
            qldYearArr.append(row[1])
        elif (var == 'South Australia'):
            saTempArr.append(row[2])
            saYearArr.append(row[1])
        elif (var == 'Tasmania'):
            tasTempArr.append(row[2])
            tasYearArr.append(row[1])
        elif (var == 'Victoria'):
            vicTempArr.append(row[2])
            vicYearArr.append(row[1])
        elif (var == 'Western Australia'):
            waTempArr.append(row[2])
            waYearArr.append(row[1])
        elif (var == 'Australia'):
            ausTempArr.append(row[2])
            ausYearArr.append(row[1])


def DataInsertFunction(_query):
    SelectStateTempData(_query)
    InsertDataIntoWorkbook(worldTemp_worksheet)


def PlotData():
    _actTempArrNP = numpy.array(actTempArr)
    _actYearArrNP = numpy.array(actYearArr)
    _nswTempArrNP = numpy.array(nswTempArr)
    _nswYearArrNP = numpy.array(nswYearArr)
    _ntTempArrNP = numpy.array(ntTempArr)
    _ntYearArrNP = numpy.array(ntYearArr)
    _qldTempArrNP = numpy.array(qldTempArr)
    _qldYearArrNP = numpy.array(qldYearArr)
    _saTempArrNP = numpy.array(saTempArr)
    _saYearArrNP = numpy.array(saYearArr)
    _tasTempArrNP = numpy.array(tasTempArr)
    _tasYearArrNP = numpy.array(tasYearArr)
    _vicTempArrNP = numpy.array(vicTempArr)
    _vicYearArrNP = numpy.array(vicYearArr)
    _waTempArrNP = numpy.array(waTempArr)
    _waYearArrNP = numpy.array(waYearArr)
    _ausTempArrNP = numpy.array(ausTempArr)
    _ausYearArrNP = numpy.array(ausYearArr)

    actLine, = pyplot.plot(_actYearArrNP, _actTempArrNP, label='ACT')
    pyplot.plot(_nswYearArrNP, _nswTempArrNP, label='NSW')
    pyplot.plot(_ntYearArrNP, _ntTempArrNP, label='NT')
    pyplot.plot(_qldYearArrNP, _qldTempArrNP, label='QLD')
    pyplot.plot(_saYearArrNP, _saTempArrNP, label='SA')
    pyplot.plot(_tasYearArrNP, _tasTempArrNP, label='TAS')
    pyplot.plot(_vicYearArrNP, _vicTempArrNP, label='VIC')
    pyplot.plot(_waYearArrNP, _waTempArrNP, label='WA')
    pyplot.plot(_ausYearArrNP, _ausTempArrNP, label='AUS')
    pyplot.title('Australian State and National Temperature Data')
    pyplot.legend(handler_map={actLine: HandlerLine2D(numpoints=4)}, bbox_to_anchor=(1, 1), loc=2)
    pyplot.ylabel('Temperature')
    pyplot.xlabel('Year')
    pyplot.axis([1841, 2013, 0, 40])
    pyplot.show()


worldTemp_workbook = LoadWorkbook(wb_filename)
worldTemp_worksheet = AddWorkbookSheet(worldTemp_workbook)

DataInsertFunction("SELECT State, SUBSTR(EventDate, 0, 5) as Year, AVG(AverageTemperature) FROM GlobalLandTemperaturesByState WHERE AverageTemperature != 'NULL' AND Country = 'Australia' GROUP BY State, Year")
DataInsertFunction("SELECT Country, SUBSTR(EventDate, 0, 5) as Year, AVG(AverageTemperature) FROM GlobalLandTemperaturesByState WHERE AverageTemperature != 'NULL' AND Country = 'Australia' GROUP BY Country, Year")
worldTemp_workbook.save(wb_filename)

ImportPlotData(worldTemp_worksheet)
PlotData()

utilities.CloseDatabaseConnection(conn)